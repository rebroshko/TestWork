import time
from openpyxl import load_workbook
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class XlData:                                                   # parsing table data in dict
    def __init__(self, table_name, list_name):
        try:
            self.wb = load_workbook(table_name)
        except Exception as exc:
            raise Exception('table name ERROR') from exc
        self.sheet_ranges = self.wb[list_name]
        self.dict_data = self.parsing_data()

    def parsing_data(self):
        self.dict_data = {}
        for row in self.sheet_ranges.iter_rows():
            a = row[0].value
            c = row[1].value
            self.dict_data[a] = c
        return self.dict_data


def parsing_url(browser):                                           # parsing and verification url
    array_xpath = create_xpath()
    local_index = array_xpath[0]
    WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.XPATH, local_index)))
    for i in array_xpath:
        local_url = browser.find_element(By.XPATH, i).get_attribute("href")
        if not local_url[8:12] == 'yabs':
            return local_url


def iterable_count(browser, value):                                 # get number half
    digit_write = '//*[@id="credit-sum"]/div[1]/input'             # string for number
    path_for_mount_half = '//*[@id="monthly-payment"]/span'
    try:
        elem_for_digit = browser.find_element(By.XPATH, digit_write)
    except Exception as exc:
        raise Exception('Search panel dont found') from exc
    for i in range(8):
        elem_for_digit.send_keys(Keys.BACKSPACE)
    elem_for_digit.send_keys(int(str(value)[0:-1]))
    time.sleep(1)
    element_to_hover_over = browser.find_element(By.XPATH, '//*[@id="bank-offer"]/button')
    hover = ActionChains(browser).move_to_element(element_to_hover_over)
    hover.perform()
    half = browser.find_element(By.XPATH, path_for_mount_half).text[0:-2]
    return half


def create_xpath():                                                 # create xpath for get url
    array_xpath_search = []
    first_path = '//*[@id="search-result"]/li'
    second_path = '/div/div[2]/div[1]/a'
    for i in range(1, 5):
        array_xpath_search.append(first_path+'['+str(i)+']'+second_path)
    return array_xpath_search


def write_data(data):                                               # Create and write in txt file data
    with open("credit.txt", "w+", encoding='UTF-8') as file:
        for key, value in data.items():
            vl = value.replace(' ', '')
            file.write(f'Сумма кредита - {key}, ежемесячный платеж - {vl}' + '\n')
